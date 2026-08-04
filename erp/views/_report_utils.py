"""
erp/views/_report_utils.py
Shared report utilities — status colours, Excel/PDF export, drill-down table.
Import from any view with:
    from ._report_utils import OP_STATUS, render_export_buttons, render_drilldown_table
"""
from __future__ import annotations

import io
from datetime import date

import pandas as pd
import streamlit as st


# ══════════════════════════════════════════════════════════════════════════════
# COLOR PALETTE  (single source of truth for the whole ERP)
# ══════════════════════════════════════════════════════════════════════════════

# Operational Status  (🟢🔵🟡🟠🔴⚫)
OP_STATUS: dict[str, tuple[str, str]] = {
    "On Rent":    ("#DCFCE7", "#166534"),   # green
    "Available":  ("#DBEAFE", "#1E40AF"),   # blue
    "Idle":       ("#DBEAFE", "#1E40AF"),   # blue  (same as Available)
    "Reserved":   ("#FEF3C7", "#92400E"),   # amber
    "In Transit": ("#FFEDD5", "#9A3412"),   # orange
    "Breakdown":  ("#FEE2E2", "#991B1B"),   # red
    "Sold":       ("#F3F4F6", "#374151"),   # dark gray
}

# Compliance / expiry status
COMPLIANCE_STATUS: dict[str, tuple[str, str]] = {
    "Valid":         ("#DCFCE7", "#166534"),
    "Expiring Soon": ("#FEF3C7", "#92400E"),
    "Overdue":       ("#FEE2E2", "#991B1B"),
    "Not Set":       ("#F1F5F9", "#9CA3AF"),
    "Invalid":       ("#F1F5F9", "#9CA3AF"),
}

# Worklog status
WL_STATUS: dict[str, tuple[str, str]] = {
    "Submitted":       ("#DCFCE7", "#166534"),
    "Draft":           ("#FEF3C7", "#92400E"),
    "Missing":         ("#FEE2E2", "#991B1B"),
    "Pending Billing": ("#FEF3C7", "#92400E"),
    "Pending":         ("#FEE2E2", "#991B1B"),
}

# Utilisation bands
def util_band(pct: float) -> tuple[str, str]:
    """(bg, fg) for a utilisation percentage — green / amber / red."""
    if pct >= 80:
        return ("#DCFCE7", "#166534")
    if pct >= 50:
        return ("#FEF3C7", "#92400E")
    return ("#FEE2E2", "#991B1B")


# ── Chip helpers ──────────────────────────────────────────────────────────────

def status_chip(label: str, bg: str = "#F1F5F9", fg: str = "#374151") -> str:
    return (
        f"<span style='background:{bg};color:{fg};padding:2px 10px;"
        f"border-radius:12px;font-size:11px;font-weight:700;"
        f"white-space:nowrap;display:inline-block;'>{label}</span>"
    )


def op_status_chip(status: str) -> str:
    bg, fg = OP_STATUS.get(status, ("#F1F5F9", "#374151"))
    return status_chip(status, bg, fg)


def compliance_chip(status: str) -> str:
    bg, fg = COMPLIANCE_STATUS.get(status, ("#F1F5F9", "#9CA3AF"))
    return status_chip(status, bg, fg)


def wl_status_chip(status: str) -> str:
    bg, fg = WL_STATUS.get(status, ("#F1F5F9", "#374151"))
    return status_chip(status, bg, fg)


# ── Standard table header style ───────────────────────────────────────────────

# Dark themed (use in overview/master tables)
TH_DARK = (
    "padding:10px 12px;"
    "background:#1E3A5F;color:#fff;"
    "font-size:10px;font-weight:700;"
    "letter-spacing:.12em;text-transform:uppercase;"
    "white-space:nowrap;"
)

# Light themed (use in nested/detail tables)
TH_LIGHT = (
    "padding:10px 12px;"
    "background:#F8FAFC;"
    "font-size:10px;font-weight:700;"
    "letter-spacing:.12em;text-transform:uppercase;color:#6B7280;"
    "border-bottom:2px solid #E2EBF0;white-space:nowrap;"
)


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT  (styled header + zebra rows)
# ══════════════════════════════════════════════════════════════════════════════

def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Report") -> bytes:
    """Return xlsx bytes with a dark styled header and zebra row fills."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        _style_ws(writer.sheets[sheet_name])
    buf.seek(0)
    return buf.getvalue()


def _style_ws(ws) -> None:
    from openpyxl.styles import PatternFill, Font, Alignment

    hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
    hdr_font  = Font(color="FFFFFF", bold=True, size=10)
    even_fill = PatternFill("solid", fgColor="F8FAFC")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")

    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.number_format = "@"          # keep header as text

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for cell in row:
            cell.fill = fill

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 46)


# ══════════════════════════════════════════════════════════════════════════════
# PDF EXPORT  (print-ready HTML — open in browser → Ctrl+P → Save as PDF)
# ══════════════════════════════════════════════════════════════════════════════

def to_pdf_html(
    title: str,
    subtitle: str,
    df: pd.DataFrame,
    extra_css: str = "",
) -> str:
    today = date.today()
    th    = "".join(f"<th>{c}</th>" for c in df.columns)
    rows  = ""
    for i, (_, row) in enumerate(df.iterrows()):
        bg    = "#FAFBFC" if i % 2 == 0 else "#FFFFFF"
        cells = "".join(
            f"<td>{'' if (v is None or (isinstance(v, float) and pd.isna(v))) else v}</td>"
            for v in row
        )
        rows += f"<tr style='background:{bg}'>{cells}</tr>"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>{title}</title>
<style>
  body  {{font-family:Arial,sans-serif;font-size:11px;margin:24px;color:#111;}}
  h1    {{font-size:20px;font-weight:800;color:#1E3A5F;margin:0 0 2px;}}
  .sub  {{font-size:11px;color:#6B7280;margin-bottom:20px;}}
  table {{border-collapse:collapse;width:100%;}}
  th    {{background:#1E3A5F;color:#fff;padding:8px 10px;text-align:left;
          font-size:9px;letter-spacing:.1em;text-transform:uppercase;}}
  td    {{padding:7px 10px;border-bottom:1px solid #E2EBF0;font-size:11px;
          vertical-align:middle;}}
  {extra_css}
  @media print {{body{{margin:0}}}}
</style>
</head>
<body>
<h1>{title}</h1>
<div class="sub">{subtitle} &middot; Generated: {today.strftime('%d %b %Y')} &middot; {len(df):,} rows</div>
<table>
<thead><tr>{th}</tr></thead>
<tbody>{rows}</tbody>
</table>
<script>window.onload=function(){{window.print();}}</script>
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT: STANDARD EXPORT BUTTON PAIR
# ══════════════════════════════════════════════════════════════════════════════

def render_export_buttons(
    df: pd.DataFrame,
    base_name: str,
    excel_key: str,
    pdf_key: str,
    title: str = "Report",
    subtitle: str = "",
    sheet_name: str = "Report",
) -> None:
    """Render Excel + PDF download buttons side by side."""
    c1, c2, _ = st.columns([1, 1, 6])
    with c1:
        st.download_button(
            "Export Excel",
            data=to_excel_bytes(df, sheet_name),
            file_name=f"{base_name}_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=excel_key,
            use_container_width=True,
        )
    with c2:
        st.download_button(
            "Export PDF",
            data=to_pdf_html(title, subtitle, df).encode("utf-8"),
            file_name=f"{base_name}_{date.today()}.html",
            mime="text/html",
            key=pdf_key,
            use_container_width=True,
            help="Opens in browser → Ctrl+P → Save as PDF",
        )


def render_print_export_buttons(
    df: pd.DataFrame,
    base_name: str,
    key_prefix: str,
    title: str = "Report",
    subtitle: str = "",
    sheet_name: str = "Report",
) -> None:
    """Render Print (inline) + Export PDF (download) + Export Excel (download)."""

    c1, c2, c3, _ = st.columns([1, 1, 1, 5])
    with c1:
        if st.button("🖨️ Print", key=f"{key_prefix}_print_btn", use_container_width=True):
            st.session_state[f"_rpe_print_{key_prefix}"] = True
    with c2:
        st.download_button(
            "⬇ Export PDF",
            data=to_pdf_html(title, subtitle, df).encode("utf-8"),
            file_name=f"{base_name}_{date.today()}.html",
            mime="text/html",
            key=f"{key_prefix}_pdf_btn",
            use_container_width=True,
            help="Opens in browser → Ctrl+P → Save as PDF",
        )
    with c3:
        st.download_button(
            "⬇ Export Excel",
            data=to_excel_bytes(df, sheet_name),
            file_name=f"{base_name}_{date.today()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{key_prefix}_xl_btn",
            use_container_width=True,
        )

    if st.session_state.pop(f"_rpe_print_{key_prefix}", False):
        st.iframe(to_pdf_html(title, subtitle, df), height=820)


# ══════════════════════════════════════════════════════════════════════════════
# STREAMLIT: SORTABLE + DRILL-DOWN TABLE
# ══════════════════════════════════════════════════════════════════════════════

def render_drilldown_table(
    df: pd.DataFrame,
    table_key: str,
    column_config: dict | None = None,
    height: int | None = None,
    style_fn=None,
) -> int | None:
    """
    Render a sortable, single-row-selectable dataframe.

    Returns the 0-based row index of the selected row, or None if nothing selected.

    Usage:
        idx = render_drilldown_table(summary_df, "my_tbl", column_config={...})
        if idx is not None:
            row = summary_df.iloc[idx]
            st.markdown(f"**Details: {row['Name']}**")
            # … render detail panel …

    Sorting is built into st.dataframe — click any column header.
    """
    kwargs: dict = dict(
        width="stretch",
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        key=table_key,
    )
    if column_config:
        kwargs["column_config"] = column_config
    if height:
        kwargs["height"] = height

    styled = df.style if style_fn is None else style_fn(df.style)
    event  = st.dataframe(styled, **kwargs)

    rows = getattr(getattr(event, "selection", None), "rows", [])
    return rows[0] if rows else None


# ── Conditional-formatting helpers for pandas Styler ─────────────────────────

def style_utilisation_col(col: pd.Series) -> list[str]:
    """Apply green/amber/red background to a Utilisation % column."""
    styles = []
    for v in col:
        try:
            pct = float(str(v).replace("%", "").strip())
        except (ValueError, TypeError):
            styles.append("")
            continue
        bg, fg = util_band(pct)
        styles.append(f"background-color:{bg};color:{fg};font-weight:700;")
    return styles


def style_op_status_col(col: pd.Series) -> list[str]:
    """Apply operational-status colours to a status column."""
    out = []
    for v in col:
        bg, fg = OP_STATUS.get(str(v), ("#F1F5F9", "#374151"))
        out.append(f"background-color:{bg};color:{fg};font-weight:700;")
    return out


def style_compliance_col(col: pd.Series) -> list[str]:
    """Apply compliance-status colours."""
    out = []
    for v in col:
        bg, fg = COMPLIANCE_STATUS.get(str(v), ("#F1F5F9", "#9CA3AF"))
        out.append(f"background-color:{bg};color:{fg};font-weight:700;")
    return out


def style_idle_days_col(col: pd.Series, threshold: int = 30) -> list[str]:
    """Highlight machines idle more than `threshold` days in red."""
    out = []
    for v in col:
        try:
            days = int(float(v or 0))
        except (ValueError, TypeError):
            out.append("")
            continue
        if days > threshold:
            out.append("background-color:#FEE2E2;color:#991B1B;font-weight:700;")
        elif days > 0:
            out.append("background-color:#FEF3C7;color:#92400E;")
        else:
            out.append("")
    return out
